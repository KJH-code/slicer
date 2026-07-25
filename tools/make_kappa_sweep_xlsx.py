"""
make_kappa_sweep_xlsx.py — 팀메 3D 표면 비교 도구용 테스트 Excel 생성기.

팀메 도구(tools/slicing_simulator.html 의 'Excel 표면 비교' 탭)가 기대하는 형식:
    x = κmax : 4행 B4, C4, D4, ...
    y = q0   : A열 A5, A6, A7, ...
    z        : B5부터의 행렬 (시트 이름 = z 항목)
    여러 파일을 겹치려면 시트 구성이 같아야 함.

내용은 난수가 아니라 실제 연구 데이터다: 팀메의 적응형 오버행 임계
    κ = −κmax · q/(q+q0),   q = (클러스터 경계길이)/√(클러스터 넓이)
를 우리 해석식(g(α)=n_z·cosα+d·n_r·sinα ≥ κ 의 최소각 닫힌형) 위에서 계산해,
(κmax, q0) 격자마다 다음을 구한 파라미터 민감도 표면이다:

    시트 selected_angle_deg : 오버행 면들의 최소 필요각(면적가중 평균, 도)
    시트 violating_area_pct : 어떤 각도(≤60°)로도 못 고치는 넓이 (전체 %)
    시트 max_angle_deg      : 가장 까다로운 면이 요구하는 각도 (도)

    python3 tools/make_kappa_sweep_xlsx.py            # 데모 3종 생성
    python3 tools/make_kappa_sweep_xlsx.py model.stl  # 내 STL 1개
"""

import sys
from pathlib import Path

import numpy as np
import trimesh
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from conical.meshio import center_on_axis            # noqa: E402
from conical.clusters import overhang_clusters       # noqa: E402
from conical.analytic import radial_normal           # noqa: E402

KMAX_VALUES = np.round(np.arange(0.1, 1.01, 0.1), 2)      # x축 (10개)
Q0_VALUES = np.round(np.arange(0.2, 3.01, 0.2), 2)        # y축 (15개)
ANGLE_CAP = 60.0                                          # 탐색 상한(도)


def cluster_boundary_length(mesh, faces):
    """클러스터 경계(클러스터 안·밖 면이 만나는 변)의 총 길이."""
    inside = np.zeros(len(mesh.faces), dtype=bool)
    inside[faces] = True
    adj = mesh.face_adjacency
    boundary = inside[adj[:, 0]] != inside[adj[:, 1]]
    edges = mesh.face_adjacency_edges[boundary]
    v = mesh.vertices
    return float(np.linalg.norm(v[edges[:, 0]] - v[edges[:, 1]], axis=1).sum())


def min_angle_for_kappa(nz, b, kappa, cap=ANGLE_CAP):
    """g(α)=nz·cosα+b·sinα ≥ κ 인 최소 α(도). 불가능하면 NaN. (b = d·n_r)"""
    theta = np.full(len(nz), np.nan)
    ok0 = nz >= kappa
    theta[ok0] = 0.0
    rest = ~ok0
    R = np.hypot(nz, b)
    with np.errstate(invalid="ignore", divide="ignore"):
        reachable = rest & (b > 1e-12) & (R > 1e-12) & (kappa / np.where(R > 0, R, 1) >= -1.0)
        ratio = np.clip(kappa / np.where(R > 0, R, 1.0), -1.0, 1.0)
        alpha = np.degrees(np.arcsin(ratio) - np.arctan2(nz, b))
    good = reachable & (alpha >= -1e-9) & (alpha <= cap + 1e-9)
    theta[good] = np.clip(alpha[good], 0.0, cap)
    return theta


def sweep_model(mesh):
    """(κmax, q0) 격자에서 세 지표 행렬을 계산."""
    mesh = center_on_axis(mesh.copy())
    mesh.merge_vertices()
    areas = mesh.area_faces
    total = float(areas.sum())
    nz_all = mesh.face_normals[:, 2]
    nr_all = radial_normal(mesh)

    clusters = overhang_clusters(mesh)
    cluster_info = []
    for faces in clusters:
        area = float(areas[faces].sum())
        blen = cluster_boundary_length(mesh, faces)
        q = blen / max(np.sqrt(area), 1e-9)
        cluster_info.append((faces, area, q))

    sel = np.zeros((len(Q0_VALUES), len(KMAX_VALUES)))
    vio = np.zeros_like(sel)
    mx = np.zeros_like(sel)

    for j, q0 in enumerate(Q0_VALUES):
        for i, kmax in enumerate(KMAX_VALUES):
            ang_sum = 0.0
            ang_area = 0.0
            bad_area = 0.0
            worst = 0.0
            for faces, area, q in cluster_info:
                kappa = -kmax * q / (q + q0)          # 팀메의 적응형 임계
                nz = nz_all[faces]
                best = None                            # 방향: 나은 쪽 채택
                for d in (+1.0, -1.0):
                    th = min_angle_for_kappa(nz, d * nr_all[faces], kappa)
                    nan_area = float(areas[faces][np.isnan(th)].sum())
                    mean_ok = np.nanmean(th) if np.isfinite(th).any() else np.inf
                    cand = (nan_area, mean_ok, th)
                    if best is None or cand[:2] < best[:2]:
                        best = cand
                nan_area, _, th = best
                okmask = np.isfinite(th)
                ang_sum += float((th[okmask] * areas[faces][okmask]).sum())
                ang_area += float(areas[faces][okmask].sum())
                bad_area += nan_area
                if okmask.any():
                    worst = max(worst, float(np.nanmax(th)))
            sel[j, i] = ang_sum / ang_area if ang_area > 0 else 0.0
            vio[j, i] = bad_area / total * 100.0
            mx[j, i] = worst
    return {"selected_angle_deg": sel, "violating_area_pct": vio,
            "max_angle_deg": mx}


def write_xlsx(path, model_name, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, z in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"model: {model_name}"
        ws["A2"] = f"sheet: {sheet_name} (κ=-κmax·q/(q+q0), 해석식 최소각)"
        ws["A4"] = "q0\\κmax"
        for i, k in enumerate(KMAX_VALUES):
            ws.cell(row=4, column=2 + i, value=float(k))       # B4, C4, ...
        for j, q0 in enumerate(Q0_VALUES):
            ws.cell(row=5 + j, column=1, value=float(q0))      # A5, A6, ...
            for i in range(len(KMAX_VALUES)):
                ws.cell(row=5 + j, column=2 + i, value=round(float(z[j, i]), 3))
    wb.save(path)
    print(f"saved: {path}")


def demo_models():
    m = {}
    m["sphere"] = trimesh.creation.icosphere(subdivisions=3, radius=10)
    funnel = trimesh.creation.cone(radius=14, height=5)
    funnel.apply_transform(trimesh.transformations.rotation_matrix(np.pi, [1, 0, 0]))
    m["funnel"] = funnel
    stem = trimesh.creation.cylinder(radius=2, height=14); stem.apply_translation([0, 0, 7])
    cap = trimesh.creation.cylinder(radius=8, height=3); cap.apply_translation([0, 0, 15.5])
    m["mushroom"] = trimesh.util.concatenate([stem, cap])
    return m


if __name__ == "__main__":
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        models = {p.stem: trimesh.load(p, force="mesh")}
    else:
        models = demo_models()
    for name, mesh in models.items():
        write_xlsx(f"kappa_sweep_{name}.xlsx", name, sweep_model(mesh))
